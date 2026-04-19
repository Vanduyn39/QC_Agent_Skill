import os
import json
import csv
import PyPDF2
from docx import Document
import openpyxl

def extract_text(file_path: str):
    if not os.path.exists(file_path):
        return f"File not found: {file_path}"
        
    text = ""
    try:
        # Convert file extension to lowercase for accurate comparison
        file_ext = file_path.lower()
        
        # 1. Process PDF file
        if file_ext.endswith('.pdf'):
            with open(file_path, 'rb') as file:
                reader = PyPDF2.PdfReader(file)
                for page in reader.pages:
                    extracted = page.extract_text()
                    if extracted:
                        text += extracted + "\n"
                        
        # 2. Process DOCX file
        elif file_ext.endswith('.docx'):
            doc = Document(file_path)
            for para in doc.paragraphs:
                text += para.text + "\n"
                
        # 3. Process JSON file
        elif file_ext.endswith('.json'):
            with open(file_path, 'r', encoding='utf-8') as file:
                data = json.load(file)
                # Format JSON into a readable string
                text = json.dumps(data, indent=4, ensure_ascii=False)
                
        # 4. Process CSV file
        elif file_ext.endswith('.csv'):
            with open(file_path, 'r', encoding='utf-8') as file:
                reader = csv.reader(file)
                for row in reader:
                    # Join columns with a separator (e.g., " | ")
                    text += " | ".join(str(cell) for cell in row if cell) + "\n"
                    
        # 5. Process Excel file (.xlsx)
        elif file_ext.endswith('.xlsx'):
            # data_only=True to get actual values instead of formulas
            wb = openpyxl.load_workbook(file_path, data_only=True)
            for sheetname in wb.sheetnames:
                sheet = wb[sheetname]
                text += f"\n--- Sheet: {sheetname} ---\n"
                for row in sheet.iter_rows(values_only=True):
                    # Convert data and replace empty cells with empty strings
                    row_data = [str(cell) if cell is not None else "" for cell in row]
                    # Only add if the row has at least 1 data cell
                    if any(cell.strip() for cell in row_data):
                        text += " | ".join(row_data) + "\n"
                        
        else:
            return "Unsupported file format. Only accepts: .pdf, .docx, .json, .csv, .xlsx"
            
        return text.strip()
        
    except Exception as e:
        return f"Error reading file: {str(e)}"
